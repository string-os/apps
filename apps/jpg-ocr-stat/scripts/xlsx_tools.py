#!/usr/bin/env python3
"""Read or write .xlsx files with openpyxl -> JSON.

  read   inspect a workbook: sheet names, dims, and the first rows of a sheet
  write  build a single-sheet workbook from a JSON rows file

The write mode mirrors what the stat page produces, so you can assemble an xlsx
from data extracted elsewhere (e.g. OCR + your own parsing).
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print(json.dumps({"success": False, "error": "Missing dependency: openpyxl. "
                      "Install: conda install -c conda-forge openpyxl"}))
    sys.exit(1)


def cmd_read(args):
    if not os.path.isfile(args.file):
        return {"success": False, "error": f"File not found: {args.file}"}
    try:
        wb = load_workbook(args.file, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"Cannot open workbook: {e}"}
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= args.rows:
            break
        rows.append([("" if c is None else c) for c in row])
    try:
        dims = ws.calculate_dimension()
    except Exception:
        dims = None
    return {"success": True, "file": args.file, "sheets": wb.sheetnames,
            "active_sheet": ws.title, "dimensions": dims,
            "preview_rows": rows}


def cmd_write(args):
    try:
        with open(args.rows) as f:
            data = json.load(f)
    except Exception as e:
        return {"success": False, "error": f"Cannot read rows JSON ({args.rows}): {e}. "
                "Expected a JSON array of arrays, e.g. [[\"filename\",\"date\"],[\"000.jpg\",\"2018-10-19\"]]"}
    if not isinstance(data, list) or not all(isinstance(r, list) for r in data):
        return {"success": False, "error": "rows JSON must be an array of arrays (rows of cells)."}
    out_dir = os.path.dirname(os.path.abspath(args.out))
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet
    for r in data:
        ws.append(r)
    wb.save(args.out)
    return {"success": True, "output": args.out, "sheet": args.sheet, "rows_written": len(data)}


def main():
    ap = argparse.ArgumentParser(description="Read/write xlsx files with openpyxl.")
    sub = ap.add_subparsers(dest="mode", required=True)

    r = sub.add_parser("read", help="Inspect a workbook")
    r.add_argument("--file", required=True)
    r.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    r.add_argument("--rows", type=int, default=20, help="Max preview rows")

    w = sub.add_parser("write", help="Write a single-sheet workbook from JSON rows")
    w.add_argument("--rows", required=True, help="Path to JSON file: array of arrays")
    w.add_argument("--out", required=True, help="Output .xlsx path")
    w.add_argument("--sheet", default="results", help="Sheet name")

    args = ap.parse_args()
    result = cmd_read(args) if args.mode == "read" else cmd_write(args)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    sys.exit(0 if result.get("success") else 1)


if __name__ == "__main__":
    main()
